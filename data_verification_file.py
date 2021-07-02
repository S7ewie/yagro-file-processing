from openpyxl import Workbook
from openpyxl.styles import PatternFill, GradientFill, Border, Side, Alignment, Protection, Font
import pandas as pd
from styles_and_what_not import YAGRO_GREEN


class DataVerificationFile:

    def __init__(self):
        self.workbook = Workbook()
        self.filename = "filename"
        self.data_types = {
            "missing_seed": {
                "question": "The following fields are missing seed information, could you provide your drilling information.",
                "input_headings": [
                    "Seed Planted",
                    "Drill Rate",
                    "Drill Date",
                    "Unit price"
                ],
                "data_headings": [
                    "Field Group",
                    "Field Name",
                    "Crop",
                    "Variety"
                ]
            },
            "missing_fert": {
                "question": "The following fields are missing fertiliser information, could you provide your fertiliser information.",
                "input_headings": [
                    "Fertiliser Product",
                    "Application Rate",
                    "Application Date",
                    "Unit price"
                ],
                "data_headings": [
                    "Field Group",
                    "Field Name",
                    "Crop",
                    "Variety"
                ]
            },
            "missing_chem": {
                "question": "The following fields are missing chemical information, could you provide your chemical information.",
                "input_headings": [
                    "Chemical Product",
                    "Application Rate",
                    "Application Date",
                    "Unit price"
                ],
                "data_headings": [
                    "Field Group",
                    "Field Name",
                    "Crop",
                    "Variety"
                ]
            },
            "missing_product_price": {
                "question": "The following products are missing a unit price, could you provide one. (Please indicate a pack quantity if the price supplied is for more than 1 L/kg).",
                "input_headings": [
                    "Unit price",
                    "Quantity of pack if applicable"
                ],
                "data_headings": [
                    "Product Name"
                ]
            }
        }
        self.style = {
            "font": Font(name='Arial', size=12, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='121212')
        }

    def add_year(self, year, data_missing, field_analysis):
        year_sheet = self.workbook.create_sheet(str(round(year)), 0)
        year_sheet["A1"] = "Data Verification"

        row = 3
        column = 3

        for item in data_missing:
            data_type = self.data_types[item["name"]]
            for heading in data_type["data_headings"]:
                year_sheet.cell(row=row, column=column, value=heading)
                column += 1
            for heading in data_type["input_headings"]:
                year_sheet.cell(row=row, column=column, value=heading)
                column += 1
            column = 1
            row += 1
            year_sheet.cell(row=row, column=column,
                            value=data_type["question"])
            column += 2

            if "fields" in item:
                for field in item["fields"]:
                    field_data = field_analysis[field]
                    year_sheet.cell(row=row, column=column,
                                    value=field_data["fgroup"])
                    column += 1
                    year_sheet.cell(row=row, column=column,
                                    value=field_data["name"])
                    column += 1
                    year_sheet.cell(row=row, column=column,
                                    value=field_data["crop"])
                    column += 1
                    year_sheet.cell(row=row, column=column,
                                    value=field_data["variety"])

                    column = 3
                    row += 1
            elif "products" in item:
                for product in item["products"]:
                    year_sheet.cell(row=row, column=column, value=product)
                    row += 1

            row += 2

        self.add_global_format(year_sheet)

    def add_problem_products(self, products):
        prob_prod_sheet = self.workbook.create_sheet("Product Check", 0)

        headings = ["Manufacturer", "Composition", "Other info"]

        prob_prod_sheet["A1"] = "Potential Problem Products"
        prob_prod_sheet["A2"] = "The following products do not have a rule associated with them or are in the database, have a check and see if any of them are ambiguous. A lot of these are likely to be fertilisers."

        row = 4
        column = 2

        for heading in headings:
            prob_prod_sheet.cell(row=row, column=column, value=heading)
            column += 1

        row = 5
        column = 1

        for product in products:
            prob_prod_sheet.cell(row=row, column=column, value=product)
            row += 1

    def add_fnames_for_checking(self, fnames):
        fname_sheet = self.workbook.create_sheet("Field Name Check", 0)

        fname_sheet["A1"] = "Field Names"
        fname_sheet["A2"] = "Could you take a look at the field names you have supplied below and indicate whether you would like any changed. The field names shown below will be how they are displayed on your platform."

        row = 4
        column = 1

        for field in fnames:
            fname_sheet.cell(row=row, column=column, value=field)
            row += 1

    def add_global_format(self, sheet):
        # this is not how it should be done but i cba to figure it out rn
        columns_to_width = ['C', 'D', 'E', 'F', "G", 'H', 'I', 'J']
        sheet.column_dimensions['A'].width = 75
        sheet.column_dimensions['B'].width = 5
        for i in columns_to_width:
            sheet.column_dimensions[i].width = 25
        for row in sheet.iter_rows(min_row=1, max_col=10):
            for cell in row:
                cell.font = self.style["font"]
                if cell.column == 1:
                    cell.alignment = Alignment(wrap_text=True)

    def save(self, name):
        filename = name + " - Data Verification.xlsx"
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = self.workbook
        writer.save()
